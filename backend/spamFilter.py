import openpyxl
import numpy as np
import pandas as pd
from cleanText import cleanString  
from sklearn.svm import LinearSVC
from sklearn.metrics import confusion_matrix, f1_score, precision_score, recall_score
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
import csv
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="bs4")

def is_spam(email_content,mode):

    def word_control(mail, word):
        lower_mail = mail.lower()
        lower_word = word.lower()
        if lower_word in lower_mail:
            return True
        else:
            return False

    black_data = pd.read_csv('datasets/black_list.csv')
    check = False
    result = ""
    for index, row in black_data.iterrows():
        word = row['Feature']
        weight = row['Weight']
    
        if word_control(email_content,word):
            check = True
            result = f"This email is spam. E-mail has been marked as spam due to containing the word {word}."



    if(check == False):
        def store():
            if(mode == "daily_mode"):
                workBookOld = openpyxl.load_workbook('datasets/daily.xlsx')
            elif(mode == "business_mode"):
                workBookOld = openpyxl.load_workbook('datasets/business.xlsx')
            elif(mode == "commercial_mode"):
                workBookOld = openpyxl.load_workbook('datasets/commercial.xlsx')
            else:
                workBookOld = openpyxl.load_workbook('datasets/DataSet.xlsx')


            dataSheetOld = workBookOld['Data set']

            xData = []  
            yData = []  

            rows = dataSheetOld.max_row  

            for i in range(2, rows+1):  
            
                if str(dataSheetOld.cell(row=i, column=2).value) != 'None':
                    xData.append(str(cleanString(dataSheetOld.cell(row=i, column=1).value)))  
                    if str(dataSheetOld.cell(row=i, column=2).value) == "1":  
                        yData.append(1)
                    else:  
                        yData.append(0)

            # NOT: Veriyi tamamen eğitmek için, xData ve yData'yı direkt olarak döndürebilirsiniz
            # Veriyi bu şekilde bölmek, test verileri elde etmek ve öğrenme algoritmasının F-skorunu hesaplamak içindir
            # xTrain, xTest, yTrain, yTest = train_test_split(xData, yData, test_size=0.1, random_state=0)
            # return xTrain, xTest, yTrain, yTest
            return xData, yData
        


        # def calcFScore(xTest, yTest, model, vectorizer):
        #     # Test verilerini vektörleştirme
        #     xTestMatrix = vectorizer.transform(xTest)
        #     yTestMatrix = np.asarray(yTest)

        #     # Model ile sınıflandırma yapma ve karışıklık matrisi oluşturma
        #     result = model.predict(xTestMatrix)
        #     matrix = confusion_matrix(yTestMatrix, result)

        #     # F-skoru, hassasiyet ve duyarlılığı hesaplama
        #     fScore = f1_score(yTestMatrix, result, pos_label=0)
        #     precision = precision_score(yTestMatrix, result, pos_label=0)
        #     recall = recall_score(yTestMatrix, result, pos_label=0)
        #     return fScore, precision, recall, matrix

        def predict(emailBody, model, vectorizer):

            featureMatrix = vectorizer.transform([cleanString(emailBody)])
            result = model.predict(featureMatrix)  
            print("Predicting...")

            if 1 in result:
                return "This email is spam."  
            else:
                return "This email is not spam."  
            
        model = LinearSVC(class_weight='balanced')

        xTrain, yTrain= store()

        vectorizer = TfidfVectorizer(stop_words='english', max_df=75)

        yTrainMatrix = np.asarray(yTrain)
        xTrainMatrix = vectorizer.fit_transform(xTrain)

        model.fit(xTrainMatrix, yTrainMatrix)


        result = predict(email_content, model, vectorizer)



        df = pd.read_csv('datasets/feature_weights.csv')

        word_list_spam = []

        if(result == "This email is spam."):
            for index, row in df.iterrows():
                word = row['Feature']
                weight = row['Weight']
            
                if weight > 0 and word_control(email_content,word):
                    print(f"E-mail has been marked as spam due to containing the word '{word}'.")
                    word_list_spam.append(word)




        def check_feature_in_list(csv_file, word_list):
            with open(csv_file, 'r') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    if row['Feature'] in word_list:
                        feature = row['Feature']
                        if feature in word_list:
                            word_list.remove(feature)
                        if feature in word_list_spam:
                            word_list_spam.remove(feature)
                        if not word_list:
                            return True
            return False
        
        white_data_file = 'datasets/white_list.csv'
        if check_feature_in_list(white_data_file, word_list_spam):
            return "This email is not spam."
        



        if(result == "This email is spam."):
            result = result + " E-mail has been marked as spam due to containing the word(s)"

            if(word_list_spam != []):
                result = result + " " + word_list_spam[0]

            if(len(word_list_spam) > 1):

                for i in range(1,len(word_list_spam)):
                    result = result + " ," + word_list_spam[i]
            
            result = result + "."
        
    return result
